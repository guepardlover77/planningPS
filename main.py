import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pytz
from datetime import timedelta
from PIL import Image, ImageTk


class CalendarFilterGUI:
    def __init__(self, root, icon_path=None):
        self.root = root
        self.root.title("Génération de planning")
        self.root.geometry("600x400")
        if icon_path and os.path.exists(icon_path):
            try:
                # Pour Windows, utiliser iconbitmap
                if os.name == 'nt':
                    if icon_path.endswith('.ico'):
                        self.root.iconbitmap(icon_path)
                    else:
                        icon = Image.open(icon_path)
                        icon_photo = ImageTk.PhotoImage(icon)
                        self.root.iconphoto(True, icon_photo)
                # Pour Linux/Mac, utiliser iconphoto
                else:
                    icon = Image.open(icon_path)
                    icon_photo = ImageTk.PhotoImage(icon)
                    self.root.iconphoto(True, icon_photo)
            except Exception as e:
                print(f"Erreur lors du chargement de l'icône : {e}")

        self.calendar_path = tk.StringVar()
        self.pairs_path = tk.StringVar()
        self.output_dir = tk.StringVar()

        main_frame = ttk.Frame(root, padding="20")
        main_frame.grid(row=0, column=0, sticky="WENS")

        style = ttk.Style()
        style.configure("Custom.TButton", padding=10)

        ttk.Label(main_frame, text="Fichier calendrier (.txt):").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.calendar_path, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(main_frame, text="Parcourir",
                   command=lambda: self.browse_file("calendar", [("Fichiers texte", "*.txt")]),
                   style="Custom.TButton").grid(row=0, column=2, padx=5)

        ttk.Label(main_frame, text="Fichier binômes PS (.txt):").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.pairs_path, width=50).grid(row=1, column=1, padx=5)
        ttk.Button(main_frame, text="Parcourir",
                   command=lambda: self.browse_file("pairs", [("Fichiers texte", "*.txt")]),
                   style="Custom.TButton").grid(row=1, column=2, padx=5)

        ttk.Label(main_frame, text="Dossier de sortie:").grid(row=3, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.output_dir, width=50).grid(row=3, column=1, padx=5)
        ttk.Button(main_frame, text="Parcourir", command=self.browse_output_directory,
                   style="Custom.TButton").grid(row=3, column=2, padx=5)

        ttk.Button(main_frame, text="Générer le planning", command=self.generate_schedule,
                   style="Custom.TButton").grid(row=4, column=0, columnspan=3, pady=20)

        self.status_label = ttk.Label(main_frame, text="")
        self.status_label.grid(row=5, column=0, columnspan=3, pady=5)

    def browse_file(self, file_type, file_types):
        filename = filedialog.askopenfilename(filetypes=file_types)
        if filename:
            if file_type == "calendar":
                self.calendar_path.set(filename)
            elif file_type == "pairs":
                self.pairs_path.set(filename)

    def browse_output_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.output_dir.set(directory)

    def format_as_table(self, ws):
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    def sort_worksheet_data(self, ws):
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            try:
                date_str = str(row[0])
                time_str = str(row[1])
                date_obj = datetime.strptime(f"{date_str} {time_str}", "%d/%m/%Y %H:%M")
                data.append((date_obj,) + row)
            except (ValueError, TypeError) as e:
                print(f"Erreur lors du traitement de la ligne {row}: {e}")
                continue

        data.sort(key=lambda x: x[0])

        if len(data) > 0:
            ws.delete_rows(2, ws.max_row)
            for row_data in data:
                ws.append([row_data[i] for i in range(1, len(row_data))])

    def merge_similar_cells(self, ws):
        for col in ['A', 'B', 'C', 'D', 'E']:
            merge_start = 2
            current_value = None

            for row in range(2, ws.max_row + 1):
                cell_value = ws[f'{col}{row}'].value
                if (cell_value != current_value or row == ws.max_row) and current_value is not None:
                    if row - merge_start > 1:
                        ws.merge_cells(f'{col}{merge_start}:{col}{row - 1}')
                        ws[f'{col}{merge_start}'].alignment = Alignment(vertical='center', horizontal='center')
                    merge_start = row

                current_value = cell_value
                if row == ws.max_row and cell_value == current_value and row - merge_start > 0:
                    ws.merge_cells(f'{col}{merge_start}:{col}{row}')
                    ws[f'{col}{merge_start}'].alignment = Alignment(vertical='center', horizontal='center')

    def generate_schedule(self):
        if not all([self.calendar_path.get(), self.pairs_path.get(), self.output_dir.get()]):
            messagebox.showerror("Erreur", "Veuillez sélectionner tous les fichiers requis et le dossier de sortie.")
            return

        try:
            output_file = os.path.join(self.output_dir.get(), 'planning.xlsx')

            wb = self.filter_calendar_file_to_excel(self.calendar_path.get(), self.pairs_path.get())
            for ws in wb.worksheets:
                self.sort_worksheet_data(ws)
                self.merge_similar_cells(ws)
                self.format_as_table(ws)
            wb.save(output_file)

            self.status_label.config(text=f"Fichier généré avec succès : {output_file}")
            messagebox.showinfo("Succès", f"Le planning a été généré avec succès !\n\nFichier : {output_file}")

        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de la génération : {str(e)}")
            self.status_label.config(text="Erreur lors de la génération du fichier")

    def filter_calendar_file_to_excel(self, input_file, PS_file):
        with open(PS_file, 'r', encoding='utf-8') as f:
            PS = [line.strip().replace('\t', ' ').strip() for line in f.readlines()]

        utc_tz = pytz.utc
        paris_tz = pytz.timezone("Europe/Paris")

        wb = Workbook()
        ws_by_week = {}
        week_order = []

        k = 0

        with open(input_file, 'r', encoding='utf-8') as f:
            in_event = False
            current_start_time = None
            current_end_time = None
            current_summary = None

            for line in f:
                line = line.strip()

                if line == 'BEGIN:VEVENT':
                    in_event = True
                    continue
                elif line == 'END:VEVENT':
                    in_event = False
                    start_datetime_utc = datetime.strptime(current_start_time, "%Y%m%dT%H%M%SZ").replace(tzinfo=utc_tz)
                    end_datetime_utc = datetime.strptime(current_end_time, "%Y%m%dT%H%M%SZ").replace(tzinfo=utc_tz)
                    start_datetime_paris = start_datetime_utc.astimezone(paris_tz)
                    end_datetime_paris = end_datetime_utc.astimezone(paris_tz)

                    current_time = start_datetime_paris
                    while current_time < end_datetime_paris:
                        next_time = min(current_time + timedelta(hours=1), end_datetime_paris)

                        start_week = current_time.date() - timedelta(days=current_time.weekday())

                        if start_week not in ws_by_week:
                            ws_name = start_week.strftime("Semaine du %d-%m-%Y")
                            ws = wb.create_sheet(title=ws_name)
                            ws_by_week[start_week] = ws
                            week_order.append(start_week)

                            ws['A1'] = 'Date'
                            ws['B1'] = 'Heure de début'
                            ws['C1'] = 'Heure de fin'
                            ws['D1'] = 'Cours'
                            ws['E1'] = 'PS'
                            for cell in ws[1]:
                                cell.font = Font(bold=True)

                        binome = PS[k % len(PS)]

                        k += 1

                        ws = ws_by_week[start_week]
                        row = ws.max_row + 1
                        ws[f'A{row}'] = current_time.date().strftime("%d/%m/%Y")
                        ws[f'B{row}'] = current_time.time().strftime("%H:%M")
                        ws[f'C{row}'] = next_time.time().strftime("%H:%M")
                        ws[f'D{row}'] = current_summary
                        ws[f'E{row}'] = binome

                        current_time = next_time
                    continue

                if in_event:
                    if line.startswith('DTSTART:'):
                        current_start_time = line.replace('DTSTART:', '').strip()
                    elif line.startswith('DTEND:'):
                        current_end_time = line.replace('DTEND:', '').strip()
                    elif line.startswith('SUMMARY:'):
                        current_summary = line.replace('SUMMARY:', '').strip()
        default_sheet = wb["Sheet"]
        if not any(default_sheet.iter_rows(min_row=2, max_row=2)):
            wb.remove(default_sheet)

        week_order.sort()
        for i, week in enumerate(week_order):
            ws = ws_by_week[week]
            wb._sheets.remove(ws)
            wb._sheets.insert(i, ws)

        stats_sheet = wb.create_sheet(title="Statistiques")
        stats_sheet['A1'] = 'Type'
        stats_sheet['B1'] = 'Nom'
        stats_sheet['C1'] = 'Nombre de passages'
        for cell in stats_sheet[1]:
            cell.font = Font(bold=True)

        return wb


def main():
    root = tk.Tk()
    icon_path = "cat_animation.ico"
    app = CalendarFilterGUI(root, icon_path)
    root.mainloop()


if __name__ == "__main__":
    main()
